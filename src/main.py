from docx import Document
from python_docx_replace import docx_replace
from pymorphy3 import MorphAnalyzer
from openpyxl import load_workbook
import os


class MainApp():
    def __init__(self, doc_name="./test_data/template.docx",
                 data_file="test_data/list_students.xlsx"):
        self.morph = MorphAnalyzer()
        self.target_text = {
            "learning_first": "",
            "curs_num": "",
            "group_name": "",
            "name_ro": "",
            "practice_place": "",
            "practice_start_fd": "",
            "practice_end_fd": "",
            "learning_sec": "",
            "name_im": "",
            "order_num": "",
            "order_date": "",
            "P_num": "",
            "specialization": "",
            "PM_name": "",
            "practice_start_day": "",
            "practice_start_month": "",
            "practice_start_year": "",
            "practice_end_day": "",
            "practice_end_month": "",
            "practice_start_year": "",
            "hours": "",
            "end_hour": "",
            "gender": ""
        }
        self.doc_name = doc_name
        self.data_file = data_file
        self.data_update()
        self.replace_text(self.students_update())
    
    # Основная функция замены текста
    def replace_text(self, students):
        for student in students:
            self.target_text["name_im"] = student
            self.target_text["name_ro"] = self.remorph_word(student)
            self.target_text["practice_place"] = students[student]
            self.target_text["learning_first"] = self.remorph_word(self.target_text["learning_first"])
            self.target_text["learning_sec"] = self.remorph_word(self.target_text["learning_sec"])
            
            # Открываем шаблон документа
            template = Document(self.doc_name)
            
            # Заменяем текст в шаблоне
            docx_replace(template, **self.target_text)
            
            # Сохраняем изменённый документ с новым названием
            self.save_doc(template)
            
            # Закрываем документ (хотя в python-docx явное закрытие не требуется, но всё равно освобождаем ресурс)
            del template
    
    def remorph_word(self, word):
        if len(word.split()) > 1:
            return self.fn_remorph(word)
        else:
            return self.learning_remorph(word)

    # Преобразование окончания слова час в зависимости от количества часов
    def remorph_hours(self, hours):
        if int(hours) % 10 == 1 and int(hours) % 100 != 11:
                return ""
        elif int(hours) % 10 in [2, 3, 4]:
                return "a"
        else:
                return "ов"
    
    # Преобразование одного слова
    def learning_remorph(self, word):
        # Склоняет отдельное слово, учитывая установленное значение пола. а потом всё по другому
        try:
            # Пробуем разобрать слово
            parsed = self.morph.parse(word)
            if parsed:
                first_variant = parsed[0]
                gender_tag = {'femn' if self.target_text.get("gender") == "femn" else 'masc'}
                if first_variant.inflect(gender_tag):
                    # Склоняем слово в нужный род
                    return first_variant.inflect(gender_tag).word.capitalize()
                else:
                    # Если не удалось изменить слово, возвращаем оригинал
                    return word
            else:
                return word
        except Exception as e:
            print(f"Ошибка при обработке слова '{word}': {e}")
            return word

    # Преобразование ФИО в дательный падеж
    def fn_remorph(self, full_name):
        # Преобразует ФИО в родительный падеж, предварительно определяя род.
        parts = full_name.split()
        result_parts = []
        gender = None
        
        # Определяем род по последней части (чаще всего это отчество)
        gender = self.determine_gender(parts[-1])
        excepted_part = self.is_excepted(parts[0])
            
        for part in parts:
            try:
                if not excepted_part:
                    # Попытка разбора слова
                    parsed = self.morph.parse(part)
                    if parsed:
                        first_variant = parsed[0]
                        if first_variant.inflect({'gent'}):
                            # Трансформируем в родительный падеж
                            result_parts.append(first_variant.inflect({'gent'}).word.capitalize())
                        else:
                            # Если не удалось трансформировать, оставляем как есть
                            result_parts.append(part)
                    else:
                        result_parts.append(part)
                else:
                    if gender == "femn":
                        result_parts.append(part[:-1] + "ой")
                    else:
                        result_parts.append(part + "а")
                        
                    excepted_part = False

            except Exception as e:
                print(f"Ошибка при обработке имени '{part}': {e}")
                result_parts.append(part)
        
        # Устанавливаем род в target_text
        self.target_text["gender"] = gender
        return ' '.join(result_parts)

    def is_excepted(self, cur_form):
        norm_form = self.morph.parse(cur_form)[0].normal_form.capitalize()
        
        if norm_form == cur_form and norm_form != None:
            return False
        else:
            return True

    def determine_gender(self, word):
        # Определяет род слова, используя pymorphy3.
        # Возвращает 'masc' или 'femn'.
        parses = self.morph.parse(word)
        if parses:
            first_parse = parses[0]
            gender = first_parse.tag.gender
            if gender in ['masc', 'femn']:
                return gender
        return None

    # Фнукция заполнения словаря practice_dates
    def upd_practice_dates(self, sheet):
        splited_start_date = sheet["F1"].value.split(".")
        splited_end_date = sheet["G1"].value.split(".")
        
        months = {
            "01": "января",
            "02": "февраля",
            "03": "марта",
            "04": "апреля",
            "05": "мая",
            "06": "июня",
            "07": "июля",
            "08": "августа",
            "09": "сентября",
            "10": "октября",
            "11": "ноября",
            "12": "декабря",
        }
        
        practice_dates = {
            "start_day":    splited_start_date[0],
            "start_month":  months[splited_start_date[1]],
            "start_year":   splited_start_date[2],
            "end_day":      splited_end_date[0],
            "end_month":    months[splited_end_date[1]],
            "end_year":     splited_end_date[2]
        }
        
        return practice_dates

    # Функция обновления данных в словаре из excel
    def data_update(self, data_file="test_data/list_students.xlsx"):
        workbook = load_workbook(filename=data_file, read_only=True)
        sheet = workbook.active
        
        practice_dates = self.upd_practice_dates(sheet)
        
        # Забираем конкретные значения из определённых ячеек
        self.target_text.update({
            "learning_first": sheet["C1"].value,
            "curs_num": sheet["D1"].value,
            "group_name": sheet["E1"].value,
            "practice_start_fd": sheet["F1"].value,
            "practice_end_fd": sheet["G1"].value,
            "learning_sec": sheet["H1"].value,
            "order_num": sheet["I1"].value,
            "order_date": sheet["J1"].value,
            "P_num": sheet["K1"].value,
            "specialization": sheet["L1"].value,
            "PM_name": sheet["M1"].value,
            "practice_start_day": practice_dates["start_day"],
            "practice_start_month": practice_dates["start_month"],
            "practice_start_year": practice_dates["start_year"][-1],
            "practice_end_day": practice_dates["end_day"],
            "practice_end_month": practice_dates["end_month"],
            "hours": sheet["N1"].value,
            "end_hour": self.remorph_hours(sheet["N1"].value)
        })
        
        workbook.close()

    # Функция обновления списка студентов из excel
    def students_update(self):
        workbook = load_workbook(filename=self.data_file, read_only=True)
        sheet = workbook.active
    
        # Словарь для хранения пар {имя студента : место практики}
        result_dict = {}
    
        # Проходим по каждой строке листа начиная с первой строки
        for row in sheet.iter_rows(min_row=1, min_col=1, max_col=2, values_only=True):
            if row[0]:  # Проверяем наличие имени студента
                student_name = str(row[0]).strip()
                practice_place = str(row[1]).strip() if row[1] else None
            
                # Добавляем пару {студент: практика} в словарь
                result_dict[student_name] = practice_place
        
        return result_dict
    
    # Функция сохранения документа
    def save_doc(self, doc):
        try:
            doc.save(f"results/{self.target_text['name_ro'].split()[0]} {self.target_text['group_name']}.docx")
        except FileNotFoundError:
            os.mkdir("results")
            doc.save(f"results/{self.target_text['name_ro'].split()[0]} {self.target_text['group_name']}.docx")

    
if __name__ == "__main__":
    import sys
    from openpyxl.utils.exceptions import InvalidFileException
    
    try:
        MainApp(sys.argv[1], sys.argv[2]) if len(sys.argv) > 1 else MainApp()
    except KeyboardInterrupt:
        print("Программа завершена пользователем, некоторые документы могут быть сохранены некорректно")
    except InvalidFileException:
        print(f"первый файл \"{sys.argv[1]}\" передан в неверном формате")
    except ValueError:
        print(f"второй файл \"{sys.argv[2]}\" передан в неверном формате")
    except Exception as e:
        print(f"Непредвиденная ошибка: {e}")
