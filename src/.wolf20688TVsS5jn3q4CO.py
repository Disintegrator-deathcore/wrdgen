from docx import Document
from python_docx_replace import docx_replace, docx_get_keys
from pymorphy3 import MorphAnalyzer
from openpyxl import load_workbook
import os


class MainApp():
    def __init__(self, doc_name="./test_data/template.docx"):
        self.morph = MorphAnalyzer()
        self.target_text = {
            "learning_first": "",
            "curs_num": "",
            "group_name": "",
            "name_ro": "",
            "practice_start_fd": "",
            "practice_end_fd": "",
            "learning_sec": "",
            "name_im": "",
            "order_num": "",
            "order_date": "",
            "P_num": "",
            "specialization": "",
            "PM_name": "",
            "practice_start_day": "",
            "practice_start_month": "",
            "practice_start_year": "",
            "practice_end_day": "",
            "practice_end_month": "",
            "practice_start_year": "",
            "hours": "",
            "end_hour": "",
            "gender": ""
        }
        self.doc_name = doc_name
        self.data_update()
        self.replace_text(self.students_update())
            
    def replace_text(self, students):
        for student in students:
            self.target_text["name_im"] = student
            self.target_text["name_ro"] = self.remorph_word(student)
            self.target_text["learning_first"] = self.remorph_word(self.target_text["learning_first"])
            self.target_text["learning_sec"] = self.remorph_word(self.target_text["learning_sec"])
            
            # Открываем шаблон документа
            template = Document(self.doc_name)
            
            # Заменяем текст в шаблоне
            docx_replace(template, **self.target_text)
            
            # Сохраняем изменённый документ с новым названием
            self.save_doc(template)
            
            # Закрываем документ (хотя в python-docx явное закрытие не требуется, но всё равно освобождаем ресурс)
            del template
    
    def remorph_word(self, word):
        if len(word.split()) > 1:
            return self.fn_remorph(word)
        else:
            return self.learning_remorph(word)
    
    # Преобразование одного слова
    def learning_remorph(self, word):
        # Склоняет отдельное слово, учитывая установленное значение пола. а потом всё по другому
        try:
            # Пробуем разобрать слово
            parsed = self.morph.parse(word)
            if parsed:
                first_variant = parsed[0]
                gender_tag = {'femn' if self.target_text.get("gender") == "femn" else 'masc'}
                if first_variant.inflect(gender_tag):
                    # Склоняем слово в нужный род
                    return first_variant.inflect(gender_tag).word.capitalize()
                else:
                    # Если не удалось изменить слово, возвращаем оригинал
                    return word
            else:
                return word
        except Exception as e:
            print(f"Ошибка при обработке слова '{word}': {e}")
            return word

    # Преобразование ФИО в дательный падеж
    def fn_remorph(self, full_name):
        # Преобразует ФИО в родительный падеж, предварительно определяя род.
        parts = full_name.split()
        result_parts = []
        gender = None
        
        # Определяем род по последней части (чаще всего это отчество)
        gender = self.determine_gender(parts[-1])
        excepted_part = self.is_excepted(parts[0])
            
        for part in parts:
            try:
                if not excepted_part:
                    # Попытка разбора слова
                    parsed = self.morph.parse(part)
                    if parsed:
                        first_variant = parsed[0]
                        if first_variant.inflect({'gent'}):
                            # Трансформируем в родительный падеж
                            result_parts.append(first_variant.inflect({'gent'}).word.capitalize())
                        else:
                            # Если не удалось трансформировать, оставляем как есть
                            result_parts.append(part)
                    else:
                        result_parts.append(part)
                else:
                    if gender == "femn":
                        result_parts.append(part[:-1] + "ой")
                    else:
                        result_parts.append(part + "а")
                        
                    excepted_part = False

            except Exception as e:
                print(f"Ошибка при обработке имени '{part}': {e}")
                result_parts.append(part)
        
        # Устанавливаем род в target_text
        self.target_text["gender"] = gender
        return ' '.join(result_parts)

    def is_excepted(self, cur_form):
        norm_form = self.morph.parse(cur_form)[0].normal_form.capitalize()
        
        if norm_form == cur_form and norm_form != None:
            return False
        else:
            return True

    def determine_gender(self, word):
        # Определяет род слова, используя pymorphy3.
        # Возвращает 'masc' или 'femn'.
        parses = self.morph.parse(word)
        if parses:
            first_parse = parses[0]
            gender = first_parse.tag.gender
            if gender in ['masc', 'femn']:
                return gender
        return None

    # Функция обновления данных в словаре из excel
    def data_update(self, data_file="test_data/list_students.xlsx"):
        workbook = load_workbook(filename=data_file, read_only=True)
        sheet = workbook.active
        
        # Забираем конкретные значения из определённых ячеек
        self.target_text.update({
            "learning_first": sheet["B1"].value,
            "curs_num": sheet["C1"].value,
            "group_name": sheet["D1"].value,
            "practice_start_fd": sheet["E1"].value,
            "practice_end_fd": sheet["F1"].value,
            "learning_sec": sheet["G1"].value,
            "order_num": sheet["H1"].value,
            "order_date": sheet["I1"].value,
            "P_num": sheet["J1"].value,
            "specialization": sheet["K1"].value,
            "PM_name": sheet["L1"].value,
            "practice_start_day": sheet["M1"].value,
            "practice_start_month": sheet["N1"].value,
            "practice_start_year": sheet["O1"].value,
            "practice_end_day": sheet["P1"].value,
            "practice_end_month": sheet["Q1"].value,
            "hours": sheet["R1"].value,
            "end_hour": sheet["S1"].value
        })
        
    # Функция обновления списка студентов из excel
    def students_update(self, data_file="test_data/list_students.xlsx"):
        workbook = load_workbook(filename=data_file, read_only=True)
        sheet = workbook.active
        return [row[0].strip() for row in sheet.iter_rows(min_row=1, max_col=1, values_only=True) if row[0]]
    
    # Функция сохранения документа
    def save_doc(self, doc):
        try:
            doc.save(f"results/{self.target_text['name_ro'].split()[0]} {self.target_text['group_name']}.docx")
        except FileNotFoundError:
            os.mkdir("results")
            doc.save(f"results/{self.target_text['name_ro'].split()[0]} {self.target_text['group_name']}.docx")

    
if __name__ == "__main__":
    try:
        MainApp()
    except KeyboardInterrupt:
        print("Программа завершена пользователем, некоторые документы могут быть сохранены некорректно")
